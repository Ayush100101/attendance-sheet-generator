from flask import Flask, render_template, request, redirect, url_for, send_file
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
PROCESSED_FOLDER = 'processed'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PROCESSED_FOLDER, exist_ok=True)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    # Get form data
    subject = request.form['subject']
    file = request.files['file']

    # Save the uploaded file
    file_path = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(file_path)

    # Process the file
    output_path = process_file(file_path, subject)

    # Provide the processed file for download
    return send_file(output_path, as_attachment=True)

@app.route('/update_subject', methods=['GET', 'POST'])
def update_subject():
    if request.method == 'POST':
        usn = request.form['usn']
        current_subject = request.form['current_subject']
        new_subject = request.form['new_subject']

        # Process the subject update
        success = update_student_subject(usn, current_subject, new_subject)

        if success:
            return "Subject updated successfully!"
        else:
            return "Error: USN not found or current subject mismatch."

    return render_template('update_subject.html')

def process_file(file_path, subject):
    # Read the Excel file
    df = pd.read_excel(file_path)

    # Debugging: Print column names to ensure proper mapping
    print("Columns in the uploaded file:", df.columns.tolist())

    # Filter students by the selected subject (search across all columns)
    filtered_df = df[df.apply(lambda row: row.astype(str).str.contains(subject, case=False, na=False).any(), axis=1)]

    # Sort students by Batch and USN for consistent ordering
    filtered_df = filtered_df.sort_values(by=['Batch', 'USN'])

    # Create a new workbook for the output
    wb = Workbook()

    # Add a sheet for the attendance template
    for batch_name, batch_data in filtered_df.groupby('Batch'):
        ws = wb.create_sheet(title=batch_name)

        # Add the subject title
        ws.merge_cells('A1:T1')
        ws['A1'] = f"Subject: {subject}"
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

        # Add the institute name
        ws.merge_cells('B2:T2')
        ws['B2'] = "Ramrao Adik Institute of Technology \n D Y Patil Deemed to be University"
        ws['B2'].alignment = Alignment(horizontal="center", vertical="center")

        # Add the headers for the attendance table
        headers = ["Sr No", "Div", "Name of Student", "USN", "Roll No"] + [str(i) for i in range(1, 20)]
        ws.append(headers)

        # Populate student details
        for i, row in enumerate(batch_data.itertuples(index=False), start=1):
            ws.append([
                i,                              # Sr No
                getattr(row, 'Division', ''),   # Division
                getattr(row, 'Name', ''),       # Name of Student
                getattr(row, 'USN', ''),        # USN
                getattr(row, 'Roll No', getattr(row, 'Roll_No', ''))  # Roll No (handles both 'Roll No' and 'Roll_No')
            ] + ["" for _ in range(19)])        # 19 empty columns

        # Adjust column widths
        for col_idx, col in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

    # Remove the default sheet created by openpyxl
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Save the generated workbook
    output_path = os.path.join(PROCESSED_FOLDER, 'Filtered_Students.xlsx')
    wb.save(output_path)

    return output_path

def update_student_subject(usn, current_subject, new_subject):
    try:
        # Locate the files for the current and new subjects
        current_file_path = os.path.join(PROCESSED_FOLDER, f"{current_subject}_Filtered_Students.xlsx")
        new_file_path = os.path.join(PROCESSED_FOLDER, f"{new_subject}_Filtered_Students.xlsx")

        if not os.path.exists(current_file_path):
            return False

        # Read the current subject file
        current_df = pd.read_excel(current_file_path)

        # Find the student by USN
        student_row = current_df[current_df['USN'].str.upper() == usn.upper()]
        if student_row.empty:
            return False

        # Remove the student from the current subject
        current_df = current_df[current_df['USN'].str.upper() != usn.upper()]
        current_df.to_excel(current_file_path, index=False)

        # Add the student to the new subject file
        if os.path.exists(new_file_path):
            new_df = pd.read_excel(new_file_path)
        else:
            new_df = pd.DataFrame(columns=current_df.columns)  # Create a new file structure if it doesn't exist

        new_df = pd.concat([new_df, student_row], ignore_index=True)
        new_df.to_excel(new_file_path, index=False)

        return True
    except Exception as e:
        print("Error updating subject:", e)
        return False

if __name__ == '__main__':
    app.run(debug=True)

    
